import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import random as rand
import math
import copy
import numpy as np
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

def normalizar(lista):
    suma = sum(lista)
    for i in range(len(lista)):
        lista[i] = lista[i]/suma
    return lista

def nivel_cliente():
    nivel = rand.randrange(1,4)
    return(nivel)

def tiempo_cliente():
    tiempo = rand.randrange(1,1000)
    return(tiempo)

# LOS SIGUIENTES DATOS (HARD CODE) UNICAMENTE SE USAN PARA GENERAR ALTAS ALEATORIAS, LOS DATOS IMPUTADOS POR EL CLIENTE SE RESPETAN

clientes = [
    "FreshMart", "EcoStore", "QuickShop", "DailyMart", "GreenWay",
    "ValueStop", "EasyGrocer", "MegaFresh", "FoodHaven", "BudgetBazaar",
    "UrbanMarket", "HappyGrocer", "PrimeFoods", "SmartMart", "GrocerTown",
    "PureFoods", "LocalBasket", "NatureNest", "QuickCart", "SaveMore",
    "FarmPick", "CityGrocer", "MarketPlace", "ChoiceMart", "GoGrocer",
    "FreshNest", "HarvestMart", "BrightBasket", "DailyChoice", "TopMart"
]


lista_nivel = []
lista_tiempo = []
for i in range(len(clientes)):
    lista_nivel.append(nivel_cliente())
    lista_tiempo.append(tiempo_cliente())
lista_nivel = normalizar(lista_nivel)
lista_tiempo = normalizar(lista_tiempo)

tabla_clientes = []
for i in range(len(clientes)):
    tabla_clientes.append([clientes[i], lista_nivel[i], lista_tiempo[i]])


dificutlad = [2,3,2,3,3,4]
precio = [60,100,50,80,80,130]
lista_dificultad = normalizar(dificutlad)
lista_precio = normalizar(precio)
tabla_productos = []
for i in range(1,7):
    tabla_productos.append(["S"+str(i), lista_dificultad[i-1], lista_precio[i-1]])

productos = [
    "S1_500ml","S2_1000ml","S3_500ml","S4_1000ml","S5_500ml","S6_1000ml",
]

cantidad_productos = [40,80,120,160,200,240,280,320,360,400]

def generar_pedido_aleatorio():

    cliente = rand.choices(clientes, k = 1)[0]
    producto_botella = rand.choices(productos, weights = [0.20,0.15,0.2,0.15,0.2,0.15], k = 1)[0]
    producto = producto_botella.split("_")[0]
    botella = producto_botella.split("_")[1]
    cantidad = rand.choices(cantidad_productos, weights = [0.5,0.3,0.1,0.03,0.02,0.01,0.01,0.01,0.01,0.01], k = 1)[0]
    return [cliente,f"{producto}*{cantidad}_{botella}"]

def valor_de_pedido(pedido,parametros,cantidad_total):
    producto = (pedido[1].split("_")[0]).split("*")[0]
    cant = int((pedido[1].split("_")[0]).split("*")[1])
    cant_norm = cant/cantidad_total
    clien = pedido[0]
    for i in range (len(tabla_productos)):
        if tabla_productos[i][0]==producto:
            dif = tabla_productos[i][1]
            prec = tabla_productos[i][2]
            break
    for i in range (len(tabla_clientes)):
        if tabla_clientes[i][0]==clien:
            niv = tabla_clientes[i][1]
            tie = tabla_clientes[i][2]
            break
    b1 = parametros[0]
    b2 = parametros[1]
    a1 = parametros[2]
    a2 = parametros[3]
    l1 = parametros[4]
    l2 = parametros[5]
    l3 = parametros[6]
    return (10*b1*(a1*niv + a2*tie) + b2*(l1*cant_norm + 2*l2*prec + l3/2*(1 -abs(dif))))

def pedidos_a_datos(pedidos, param):
    cantidad_total = 0
    for pedido in pedidos:
        cantidad_total += int((pedido[1].split("_")[0]).split("*")[1])
    for pedido in pedidos:

        valor = valor_de_pedido(pedido, param, cantidad_total)

        pedido.insert(0,valor)

    pedidos_ordenados = sorted(pedidos, reverse = True)

    return pedidos_ordenados

import copy

import copy

# Modificación de la función agregar_pedido_forzoso
def agregar_pedido_forzoso(maquinas, pedido_forzoso, waitlist, descompuestas, sim_pedidos_dict, max_intentos=10):
    intentos = 0
    pedidos_bloqueados = set()

    while pedido_forzoso in waitlist and intentos < max_intentos:
        maquinas, waitlist = agregar_pedido(maquinas, pedido_forzoso, waitlist, descompuestas, sim_pedidos_dict)
        
        if pedido_forzoso not in waitlist:
            # Pedido agregado con éxito, salir del bucle
            break

        # Si no se logró agregar, eliminamos el pedido más reciente realizado y tratamos de hacer espacio
        pedidos_realizados, _ = pedidos_realizados_y_no_realizados(maquinas, waitlist, sim_pedidos_dict)
        if pedidos_realizados:
            pedido_a_eliminar = pedidos_realizados[-1]
            if pedido_a_eliminar not in waitlist:
                waitlist.append(sim_pedidos_dict[pedido_a_eliminar])
            maquinas, waitlist = eliminar_pedido(maquinas, pedido_a_eliminar, waitlist, sim_pedidos_dict, descompuestas)

        # Incrementamos el contador de intentos
        intentos += 1

    # Si se alcanzó el límite de intentos y aún no se pudo agregar
    if pedido_forzoso in waitlist:
        print(f"No se pudo agregar el pedido forzoso después de {max_intentos} intentos.")
        pedidos_bloqueados.add(pedido_forzoso)

    # Intentamos agregar los pedidos restantes en la lista de espera (sin forzar)
    try_wait_list = copy.deepcopy(waitlist)
    for pedido in try_wait_list:
        if pedido not in pedidos_bloqueados:
            maquinas, waitlist = agregar_pedido(maquinas, pedido, waitlist, descompuestas, sim_pedidos_dict)

    return maquinas, waitlist

# Modificación de la función eliminar_pedido
def eliminar_pedido(maquinas, pedido, waitlist, sim_pedidos_dict, descompuestas, max_intentos=10):
    # Eliminamos el pedido de las máquinas
    for i in range(len(maquinas)):
        maquinas[i] = [" " if item == pedido else item for item in maquinas[i]]

    # Intentamos re-agregar los pedidos que quedaron en la lista de espera
    try_wait_list = copy.deepcopy(waitlist)
    intentos = 0
    pedidos_bloqueados = set()

    while try_wait_list and intentos < max_intentos:
        for pedido in try_wait_list:
            if pedido not in pedidos_bloqueados:
                maquinas, waitlist = agregar_pedido(maquinas, pedido, waitlist, descompuestas, sim_pedidos_dict)

        # Actualizamos `try_wait_list` para reflejar cualquier cambio en `waitlist`
        try_wait_list = [p for p in waitlist if p in try_wait_list and p not in pedidos_bloqueados]

        # Incrementamos el contador de intentos
        intentos += 1

    # Si alcanzamos el número máximo de intentos y todavía hay pedidos no re-agregados
    if try_wait_list:
        print(f"No se pudo re-agregar algunos pedidos después de {max_intentos} intentos.")
        pedidos_bloqueados.update(try_wait_list)

    return maquinas, waitlist


# FUNCION 2
def descomponer_maquina(maquinas, maq, descompuestas, waitlist, sim_pedidos_dict):
    if maq not in descompuestas:
        descompuestas.append(maq)
    maq = maq[0]
    pedidos_a_reprogramar = list(set(maquinas[maq-1]) - {' '})
    if pedidos_a_reprogramar:
        for p in pedidos_a_reprogramar:
            maquinas, waitlist = eliminar_pedido(maquinas, p, waitlist, sim_pedidos_dict, descompuestas)
    return maquinas, descompuestas, pedidos_a_reprogramar, waitlist


# Funciones Generales

def maquinas_posibles(_pedido):
    tam = _pedido.split("_")[1]
    if tam == "500ml":
        return [3,5]
    else:
        return [4,5]

def litros_pedido(_pedido):
    listz = (_pedido.split("*")[1]).split("_")
    litros = int(listz[0])*int(listz[1].split("ml")[0])/1000
    return int(litros)

def datos_pedido(_pedido):
    posibles = maquinas_posibles(_pedido)
    litros = litros_pedido(_pedido)
    return posibles, litros

def pedidos_realizados_y_no_realizados(maquinas,waitlist,pedidos_dict):
    pedidos_realizados = sorted(set([pedido for sublist in maquinas for pedido in sublist if pedido.strip()]))
    pedidos_no_realizados_previo = [i for i in waitlist]
    pedidos_no_realizados = [list(filter(lambda y: pedidos_dict[y] == x, pedidos_dict))[0] for x in waitlist]
    return pedidos_realizados, pedidos_no_realizados

def slots_subfunction(maquina, slots, minimum_start):
    blank_counter = 0
    for idx in range(minimum_start, len(maquina)): 
        if maquina[idx] == " ":
            blank_counter += 1
            if blank_counter == slots:
                return idx - slots + 1, idx
        else:
            blank_counter = 0 

    return False, False

def empty_slots_bottle(maquina, cook_id, bottle_id, litros, prod, end_of_cook):
    check_id_list = []  
    batch_size = 80 if cook_id == 2 else 40  
    cook_length = 5 if cook_id == 2 else 4 

    if bottle_id == 3:
        bottle_capacity = 40
    elif bottle_id == 4:
        bottle_capacity = 80
    elif bottle_id == 5:
        if prod in ["P1","P3","P5"]:
            bottle_capacity = 40
        else:
            bottle_capacity = 80

    remaining_litros = litros
    current_cooking_end = end_of_cook - (math.ceil(litros / batch_size) - 1) * cook_length  

    while remaining_litros > 0:
        batch_litros = min(batch_size, remaining_litros) 
        remaining_litros -= batch_litros

        bottle_process_start = current_cooking_end + 1

        while batch_litros > 0:
            process_litros = min(bottle_capacity, batch_litros)
            batch_litros -= process_litros

            process_start = bottle_process_start
            if bottle_id == 3:
                step = 1
            elif bottle_id == 4:
                step = 2
            elif bottle_id == 5:
                if prod in ["P1","P3","P5"]:
                    step = 2
                else:
                    step = 3
            process_end = process_start + step

            check_id_list.extend(range(process_start, process_end + 1)) 

            bottle_process_start = process_end + 1

        current_cooking_end += cook_length

    for idx in check_id_list:
        try:
            if maquina[idx] != " ":
                return False
        except:
            return False
    return check_id_list


def empty_slots_cook(maquina, maquina_id, litros, minimum=0):
    blank_counter = 0

    if maquina_id == 0 or maquina_id == 1:
        slots = max(4, 4 * math.ceil(litros / 40))
        start_id, end_id = slots_subfunction(maquina, slots, minimum)
        if start_id is not False:
            return start_id, end_id

    if maquina_id == 2:
        slots = max(5, 5 * math.ceil(litros / 80))
        start_id, end_id = slots_subfunction(maquina, slots, minimum)
        if start_id is not False:
            return start_id, end_id

    return False, False

def try_add(maquinas, posibles, litros, nombre, prod, descompuestas):
    minimum_start = 0 

    while minimum_start <= 47:
        cook_start_ids = []
        cook_end_ids = []

        for i in range(3):
            
            if i+1 in descompuestas:
                start_id, end_id = False, False
            else:  
                start_id, end_id = empty_slots_cook(maquinas[i], i, litros, minimum_start)
            cook_start_ids.append(start_id)
            cook_end_ids.append(end_id)
        
        if all(start_id is False for start_id in cook_start_ids):
            #print("No exiten espacios disponibles para semiterminado.")
            return maquinas

        tried_bottles = False
        while any(start_id is not False for start_id in cook_start_ids):
            min_cook_end_index = min(
                (idx for idx, end in enumerate(cook_end_ids) if end is not False),
                key=lambda idx: cook_end_ids[idx]
            )
            min_cook_start_id = cook_start_ids[min_cook_end_index]
            min_cook_end_id = cook_end_ids[min_cook_end_index]
            cook_machine = min_cook_end_index
            
            bottle_processes = []
            bottle_end_ids = []
            
            for j in posibles:
                if j in descompuestas:
                    end_id = False
                else:
                    bottle_id_list = empty_slots_bottle(maquinas[j], cook_machine, j, litros, prod, min_cook_end_id)
                    try:
                        end_id = bottle_id_list[-1]
                    except:
                        end_id = False
                    bottle_processes.append(bottle_id_list)
                    bottle_end_ids.append(end_id)

            if all(end_id is False for end_id in bottle_end_ids):
                cook_start_ids[min_cook_end_index] = False
                cook_end_ids[min_cook_end_index] = False
                if tried_bottles:
                    break
                tried_bottles = True
                continue

            min_bottle_end_index = min(
                (idx for idx, end in enumerate(bottle_end_ids) if end is not False),
                key=lambda idx: bottle_end_ids[idx]
            )
            
            min_bottle_end_id = bottle_end_ids[min_bottle_end_index]
            bottle_machine = posibles[min_bottle_end_index]
            final_bottle_ids = bottle_processes[min_bottle_end_index]
            
            print(f"Semiterminado: M{cook_machine+1} ({min_cook_start_id}, {min_cook_end_id})")
            print(f"Envasado: M{bottle_machine+1} ({final_bottle_ids})")

            for i in range(min_cook_start_id, min_cook_end_id + 1):
                maquinas[cook_machine][i] = nombre

            for j in final_bottle_ids:
                maquinas[bottle_machine][j] = nombre

            return maquinas 

        minimum_start += 1

    #print("No existe espacio para agregar el pedido.")
    return maquinas


def agregar_pedido(maquinas, pedido, waitlist, descompuestas, pedidos_dict):
    try:
        maquina_og = copy.deepcopy(maquinas)
        maquina_previo = copy.deepcopy(maquinas)
        posibles, litros = datos_pedido(pedido[2])
        codigo_pedido = list(filter(lambda x: pedidos_dict[x] == pedido, pedidos_dict))[0]
        print(codigo_pedido, pedido)
        print(litros, "Litros")
        prod = pedido[2].split("*")[0]
        maquina_post = try_add(maquina_previo, posibles, litros, codigo_pedido, prod, descompuestas)
        
        if maquina_post == maquina_og:
            print(f"\nNo se logró agregar {pedido} ({codigo_pedido})\n")
            if pedido not in waitlist:
                waitlist.append(pedido)
        else:
            if pedido in waitlist:
                waitlist.remove(pedido)

        return maquina_post, waitlist

    except Exception as e:
        print(f"Ocurrió un error al agregar el pedido: {str(e)}")
        # Retornamos los valores originales para no perder el estado de las máquinas y la lista de espera
        return maquinas, waitlist


def generar_pedido_forzoso(running_pedidos, params):
    running_pedidos += 1
    nuevo_pedido = generar_pedido_aleatorio()
    nuevo_pedido = pedidos_a_datos([nuevo_pedido], params)
    nuevo_pedido[0][0] = 99999
    nuevo_pedido = nuevo_pedido[0]
    return nuevo_pedido, running_pedidos

def seleccionar_archivo():
    file_path = filedialog.askopenfilename(
        title="Seleccione un archivo Excel",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if file_path:
        file_entry.delete(0, tk.END)
        file_entry.insert(0, file_path)
        try:
            maquinas_df = pd.read_excel(file_path, sheet_name="Hoja1")
            pedidos_df = pd.read_excel(file_path, sheet_name="Hoja2")
            return maquinas_df, pedidos_df
        except Exception as e:
            messagebox.showerror("Error", f"Error al leer el archivo: {str(e)}")
            return None, None
    else:
        messagebox.showwarning("Advertencia", "No se seleccionó ningún archivo.")
        return None, None
    
def orden_de_reprogramacion(n_altas, n_bajas, n_maquinas):
    lista_reprogramacion = []
    
    for i in range(n_altas):
        lista_reprogramacion.append("A")
    for j in range(n_bajas):
        lista_reprogramacion.append("B")
    for k in range(n_maquinas):
        lista_reprogramacion.append("M")
        
    rand.shuffle(lista_reprogramacion)
    
    return lista_reprogramacion
    
def generar_eventos(media_altas, media_bajas, media_descomposiciones):
    num_altas = np.random.poisson(media_altas)
    num_bajas = np.random.poisson(media_bajas)
    num_descomposiciones = min(1, np.random.poisson(media_descomposiciones))

    return num_altas, num_bajas, num_descomposiciones

def procesar_diccionario_y_generar_excel(diccionario):
    filas = []

    for clave, valor in diccionario.items():
        pedido = clave
        tienda = valor[1] 
        producto_info = valor[2]  
        
        match = re.match(r"(.+)\*(.+)_(.+)", producto_info)
        if match:
            producto, demanda, tamanio = match.groups()
            producto = producto.replace('P', 'S')
            filas.append([pedido, tienda, producto, demanda, tamanio])

    df = pd.DataFrame(filas, columns=["Pedido", "Tienda", "Producto", "Demanda", "Tamaño"])
    
    return df

def crear_excel_bonito(caso_inicial):

    columns = []
    for hour in range(8, 16):
        for minute in range(0, 60, 10):
            columns.append(f"{hour:02d}:{minute:02d}")

    my_index = ["M1", "M2", "M3", "E1", "E2", "E3"]

    df = pd.DataFrame(caso_inicial, columns=columns, index=my_index)

    temp_file = "temp.xlsx"
    df.to_excel(temp_file)

    wb = load_workbook(temp_file)
    ws = wb.active
    ws.title = "Programacion"

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            if cell.value and cell.value.strip(): 
                cell.value = f"Pedido {cell.value}"
                cell.alignment = Alignment(horizontal='center', vertical='center')
                letter = cell.value.split()[-1]  
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
    return wb
 
def correr_simulacion(sim_num, maquinas, sim_pedidos_dict, running_pedidos, params, lista_params_modelo):
    
    n_altas, n_bajas, n_maquinas = params
    lista_reprogramacion = orden_de_reprogramacion(n_altas, n_bajas, n_maquinas)
    
    wait_list = []
    descompuestas = []
    
    for reprogramacion in lista_reprogramacion:
        
        if reprogramacion == "A":
            
            pedido_que_queremos_agregar, running_pedidos = generar_pedido_forzoso(running_pedidos, lista_params_modelo)
            sim_pedidos_dict[f"P{running_pedidos}"] = pedido_que_queremos_agregar
            maquinas, wait_list = agregar_pedido_forzoso(maquinas, pedido_que_queremos_agregar, wait_list, descompuestas, sim_pedidos_dict)
        
        if reprogramacion == "B":
             
            pedidos_realizados, pedidos_no_realizados = pedidos_realizados_y_no_realizados(maquinas,wait_list, sim_pedidos_dict)
            pedido_a_eliminar = rand.choices(pedidos_realizados, k = 1)  
            maquinas, wait_list = eliminar_pedido(maquinas, pedido_a_eliminar, wait_list, sim_pedidos_dict, descompuestas)
            
        if reprogramacion == "M":
             
             while True:
                 id_maquina_a_eliminar = rand.choices(list(range(1,7)), k=1)
                 if id_maquina_a_eliminar not in descompuestas:
                     break
                    
             maquinas, descompuestas, pedidos_a_reprogramar, wait_list = descomponer_maquina(maquinas, id_maquina_a_eliminar, descompuestas, wait_list, sim_pedidos_dict)
             
             for key in pedidos_a_reprogramar:
                pedido = sim_pedidos_dict[key]
                maquinas, wait_list = agregar_pedido(maquinas, pedido, wait_list, descompuestas, sim_pedidos_dict)
    
    return maquinas, sim_pedidos_dict
    
def validar_entrada(valor, nombre_variable,inf,sup):
    try:
        valor = int(valor)
        if valor < round(inf) or valor > sup:
            raise ValueError(f"{nombre_variable} debe estar en el rango de {inf} a {sup}.")
        return valor
    except ValueError as e:
        messagebox.showerror("Error de validación", str(e))
    
def correr_simulaciones():
    file_path = file_entry.get()
    if not file_path:
        messagebox.showwarning("Advertencia", "Debe seleccionar un archivo.")
        return

    maquinas_para_excel = []
    diccionarios_para_excel = []
    maquinas_para_excel_previo = []
    diccionarios_para_excel_previo = []
    
    maquinas_df, pedidos_df = seleccionar_archivo()
    if maquinas_df is None or pedidos_df is None:
        return

    listas = maquinas_df.iloc[:, 1:].values.tolist()
    maquinas_iniciales = [[valor if not pd.isna(valor) else " " for valor in sublista] for sublista in listas]  

    listas_pedidos = pedidos_df.iloc[:, 1:].values.tolist()
    listas_pedidos_final = [
        [sublista[0], f"{sublista[1]}*{sublista[2]}_{sublista[3]}ml"]
        for sublista in listas_pedidos
    ]
    

    
    altas = validar_entrada(altas_entry.get(), "Altas",0,math.inf)
    bajas = validar_entrada(bajas_entry.get(), "Bajas",0,math.inf)
    maquinas = validar_entrada(maquinas_entry.get(), "Máquinas",0,6)
    beta1 = validar_entrada(beta0_entry.get(), "Valor del Cliente",0,10)
    beta2 = validar_entrada(beta1_entry.get(), "Valor del Producto",0,10)
    alpha1 = validar_entrada(beta2_entry.get(), "Nivel del Cliente",0,10)
    alpha2 = validar_entrada(beta3_entry.get(), "Tiempo del Cliente",0,10)
    lambda1 = validar_entrada(beta4_entry.get(), "Cantidad de Producto",0,10)
    lambda2 = validar_entrada(beta5_entry.get(), "Precio del Producto",0,10)
    lambda3 = validar_entrada(beta7_entry.get(), "Dificultad del Producto",0,10)

    
    lista_params_modelo = [beta1, beta2, alpha1, alpha2, lambda1, lambda2, lambda3]
    
    pedidos_a_simular = pedidos_a_datos(listas_pedidos_final, lista_params_modelo)
    
    pedidos_dict = {}
    
    running_pedidos = 0

    for idx, pedido in enumerate(pedidos_a_simular):
        running_pedidos +=1
        pedidos_dict[f"P{running_pedidos}"] = pedido

    n_simulaciones = int(simulaciones_entry.get())

    for sim_num in range(n_simulaciones):
        pedido_dict_copia = copy.deepcopy(pedidos_dict)
        num_altas, num_bajas, num_descomposiciones = generar_eventos(altas, bajas, maquinas)
        print(num_altas, num_bajas, num_descomposiciones)
        maquinas_sim, sim_dict = correr_simulacion(sim_num, maquinas_iniciales, pedido_dict_copia, running_pedidos, [num_altas,num_bajas,num_descomposiciones], lista_params_modelo)
        maquinas_para_excel_previo.append(maquinas_sim)
        diccionarios_para_excel_previo.append(sim_dict)
    
    maquinas_workbooks = []
    diccionarios_dfs = []

    for i in range(len(maquinas_para_excel_previo)):
        maquina_wb = crear_excel_bonito(maquinas_para_excel_previo[i])
        maquinas_workbooks.append(maquina_wb)

        dict_df = procesar_diccionario_y_generar_excel(diccionarios_para_excel_previo[i])
        diccionarios_dfs.append(dict_df)

    output_file = r"C:\Users\emili\Downloads\Simulaciones_Reprogramacion_Equipo5.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        for idx, wb in enumerate(maquinas_workbooks):
            sheet = wb.active  # Get the first sheet
            data = [[cell.value for cell in row] for row in sheet.iter_rows()]
            pd.DataFrame(data).to_excel(writer, sheet_name=f'Sim{idx+1}', header=False, index=False)
        
        for idx, dict_df in enumerate(diccionarios_dfs):
            dict_df.to_excel(writer, sheet_name=f'Dict{idx+1}', index=False)

    messagebox.showinfo(
        "Éxito", 
        f"Simulación completada con éxito.\nArchivo: Simulaciones_Reprogramacion_Equipo5.xlsx"
    )

root = tk.Tk()
root.title("Simulación de Trabajos")

file_label = tk.Label(root, text="Archivo Excel:")
file_label.pack(pady=5)

file_entry = tk.Entry(root, width=40)
file_entry.pack(pady=5)

file_button = tk.Button(root, text="Seleccionar Archivo", command=seleccionar_archivo)
file_button.pack(pady=5)

altas_label = tk.Label(root, text="Media Diaria de Altas (Poisson):")
altas_label.pack(pady=5)

altas_entry = tk.Entry(root)
altas_entry.pack(pady=5)

bajas_label = tk.Label(root, text="Media Diaria de Bajas (Poisson):")
bajas_label.pack(pady=5)

bajas_entry = tk.Entry(root)
bajas_entry.pack(pady=5)

maquinas_label = tk.Label(root, text="Media Diaria de Maquinas (Poisson):")
maquinas_label.pack(pady=5)

maquinas_entry = tk.Entry(root)
maquinas_entry.pack(pady=5)

simulaciones_label = tk.Label(root, text="Cantidad Simulaciones:")
simulaciones_label.pack(pady=5)

simulaciones_entry = tk.Entry(root)
simulaciones_entry.pack(pady=5)

beta0_label = tk.Label(root, text="Peso de la función de valor de cliente (entero):")
beta0_label.pack(pady=5)

beta0_entry = tk.Entry(root)
beta0_entry.pack(pady=5)

beta1_label = tk.Label(root, text="Peso de la función de valor de producto (entero):")
beta1_label.pack(pady=5)

beta1_entry = tk.Entry(root)
beta1_entry.pack(pady=5)

beta2_label = tk.Label(root, text="Peso del nivel del cliente (entero):")
beta2_label.pack(pady=5)

beta2_entry = tk.Entry(root)
beta2_entry.pack(pady=5)

beta3_label = tk.Label(root, text="Peso del tiempo del cliente (entero):")
beta3_label.pack(pady=5)

beta3_entry = tk.Entry(root)
beta3_entry.pack(pady=5)

beta4_label = tk.Label(root, text="Peso de la cantidad de producto (entero):")
beta4_label.pack(pady=5)

beta4_entry = tk.Entry(root)
beta4_entry.pack(pady=5)

beta5_label = tk.Label(root, text="Peso del precio del producto (entero):")
beta5_label.pack(pady=5)

beta5_entry = tk.Entry(root)
beta5_entry.pack(pady=5)

beta7_label = tk.Label(root, text="Peso de la dificultad del producto (entero):")
beta7_label.pack(pady=5)

beta7_entry = tk.Entry(root)
beta7_entry.pack(pady=5)

run_button = tk.Button(root, text="Correr Simulaciones", command=correr_simulaciones)
run_button.pack(pady=10)

root.mainloop()
